#! python3
# sending due membership

import openpyxl
import smtplib
import getpass

wb = openpyxl.load_workbook('files/duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
lastMonth = sheet.cell(row=1, column=lastCol).value

unpaid = {}
# find unpaid member
for i in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=i, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=i, column=1).value
        email = sheet.cell(row=i, column=2).value
        unpaid[name] = email

# log into mail
if unpaid != {}:
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.starttls()
    print('Email: ', end='')
    my_email = input()
    smtp.login(my_email, getpass.getpass())

    for name, email in unpaid.items():
        body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for % s.Please make this " \
               "payment as soon as possible.Thank you!'" % (lastMonth, name, lastMonth)
        print('Sending email to', email)
        status = smtp.sendmail(my_email, email, body)
        if status != {}:
            print('There was a problem sending mail to %s: %s' % (email, status))
else:
    print('No Record to send')
