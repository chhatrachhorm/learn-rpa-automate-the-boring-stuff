#! python3
# program to change the miss-price per unit of
# Celery, Garlic and Lemon
# A - Product
# B - Cost Per Pound
# C - Pounds sold
# D - Total excel function => (B * C, 2)

import openpyxl

wb = openpyxl.load_workbook('../files/produceSales.xlsx')
currentSheet = wb.get_sheet_by_name('Sheet')
NEW_PRICE = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
for rowNum in range(2, currentSheet.max_row):
    productName = currentSheet.cell(row=rowNum, column=1).value

    # if the productName exist in keys in NEW_PRICE
    if productName in NEW_PRICE:
        currentSheet.cell(row=rowNum, column=2).value = NEW_PRICE[productName]
wb.save('updatedPrice.xlsx')

