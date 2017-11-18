from openpyxl.styles import Font
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')  # default sheet

# setting font - font obj
italicBold24 = Font(name='Times New Roman', size=24, italic=True, bold=True)
sheet['A1'].font = italicBold24
sheet['A1'].value = 'Hello World'  # sheet['A1]='Hello World'

# adjust rows and columns
sheet['B2'] = 'Tall Row'
sheet['C3'] = 'Wide Column'

sheet.row_dimensions[2].height = 70
sheet.column_dimensions['C'].width = 50

# merging and unmerging cells
sheet.merge_cells('D2:E9')
sheet['D2'] = 'Merged Cell'

sheet.merge_cells('D5:F6')
sheet.unmerge_cells('D5:F6')

# Freezing Panes
# 'A1' or None => UnFrozen all panes
# 'A2' => Frozen row 1
# 'B1' => Frozen Column A
# 'C1' => Frozen Column A and B
# 'C2' => Frozen Row 1 and Columns A B

sheet.freeze_panes = 'A2'

wb.save('files/custom.xlsx')
